import requests
import re
import openpyxl
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO


def sort_ammo_by_kills(workbook):
    ammo_sheet = workbook['Ammo Stats']
    
    # Получаем данные из листа Ammo Stats и сортируем их по убыванию количества убийств (колонка B)
    data = []
    for row in ammo_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        row_data = [cell if hasattr(cell, 'value') else cell for cell in row]
        data.append(tuple(row_data))

    sorted_data = sorted(data, key=lambda x: x[1], reverse=True)
    
    # Заменяем существующие данные в листе Ammo Stats отсортированными данными
    for row_number, row_data in enumerate(sorted_data, start=2):
        for col_number, value in enumerate(row_data, start=1):
            ammo_sheet.cell(row=row_number, column=col_number, value=value)


def add_ammo_types_sheet(workbook, stats_dict):
    ammo_sheet = workbook.create_sheet(title="Ammo Stats")
    ammo_sheet['A1'] = "Caliber"
    ammo_sheet['B1'] = "Kills"

    # Извлекаем данные о пулях и их убийствах
    ammo_info = [key.replace("kills-from-bullet-", "") for key in stats_dict.keys() if "kills-from-bullet-" in key]

    for i, ammo in enumerate(ammo_info, start=2):
        ammo_sheet[f'A{i}'] = ammo
        ammo_sheet[f'B{i}'] = stats_dict.get(f"kills-from-bullet-{ammo}", 0)


def clean_gun_names(workbook):
    guns_sheet = workbook['Guns']

    for row_number, row in enumerate(guns_sheet.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
        gun_name = row[0]

        # Удаляем "knife-" и "*-"
        cleaned_gun_name = re.sub(r'knife-|(\w+-)', '', gun_name)

        # Удаляем все "-" внутри строки, если они не находятся в начале или конце
        cleaned_gun_name = cleaned_gun_name.replace('-', '')
        
        guns_sheet[f'A{row_number}'] = cleaned_gun_name


def clean_ammo_names(workbook):
    ammo_sheet = workbook['Ammo Stats']

    for row_number, row in enumerate(ammo_sheet.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        ammo_caliber = row[0]

        # Удаляем "bullet-"
        cleaned_ammo_caliber = ammo_caliber.replace('bullet-', '')
        # cleaned_ammo_caliber = re.sub(r'bullet-|(\w+-)', '', gun_name)

        ammo_sheet[f'A{row_number}'] = cleaned_ammo_caliber


def sort_guns_by_kills(workbook):
    guns_sheet = workbook['Guns']
    
    # Получаем данные из листа Guns и сортируем их по убыванию количества убийств (колонка B)
    data = []
    for row in guns_sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        row_data = [cell if hasattr(cell, 'value') else cell for cell in row]
        data.append(tuple(row_data))

    sorted_data = sorted(data, key=lambda x: x[1], reverse=True)
    
    # Заменяем существующие данные в листе Guns отсортированными данными
    for row_number, row_data in enumerate(sorted_data, start=2):
        for col_number, value in enumerate(row_data, start=1):
            guns_sheet.cell(row=row_number, column=col_number, value=value)

def add_filters_to_guns(workbook):
    guns_sheet = workbook['Guns']
    
    # Добавляем заголовки для столбцов Type и Class в фильтры
    guns_sheet.auto_filter.ref = guns_sheet.dimensions
    guns_sheet['F1'] = 'Type'
    guns_sheet['G1'] = 'Class'





def get_player_stats(player_id):
    url = f"https://brainout.org/user/{player_id}"
    response = requests.get(url)

    if response.status_code == 200:
        # Ищем имя игрока в HTML-коде страницы
        name_pattern = re.compile(r'<p class="text-success"> (.*?) </p><br>', re.DOTALL)
        name_match = name_pattern.search(response.text)
        
        if name_match:
            player_name = name_match.group(1).strip()
            print(f"Player's name: {player_name}")
        else:
            print("The player's name could not be found.")
            return None, None

        # Ищем ссылку на картинку
        image_pattern = re.compile(r'<img src="(https://avatars.steamstatic.com/.*?)"', re.DOTALL)
        image_match = image_pattern.search(response.text)
        
        if image_match:
            player_image_url = image_match.group(1).strip()
            print(f"Player's image URL: {player_image_url}")
        else:
            print("The player's image could not be found.")
            return None, None

        script_pattern = re.compile(r'USER_STATS = ({.*?});', re.DOTALL)
        match = script_pattern.search(response.text)

        if match:
            player_stats_str = match.group(1)
            return player_name, player_stats_str, player_image_url
        else:
            print("The player's statistics could not be found.")
            return None, None, None
    else:
        print(f"Error when requesting a page: {response.status_code}")
        return None, None, None

def parse_stats(player_stats_str):
    stats_dict = json.loads(player_stats_str)
    return stats_dict

def create_excel_table(stats_dict, player_name, player_image_url, excel_filename):
    workbook = Workbook()

    # Главный лист
    general_sheet = workbook.active
    general_sheet.title = "General"
    general_sheet['A1'] = "Parameter"
    general_sheet['B1'] = "Value"

    # Добавляем "id" и "name" на лист "General"
    general_sheet['E3'] = "id"
    general_sheet['E4'] = "name"
    general_sheet['F3'] = player_id
    general_sheet['F4'] = player_name
    
    # cell_I4 = general_sheet['I4']
    # cell_I4.value = player_name

    # # Align the content in cell I4 to the right
    # cell_I4.alignment = openpyxl.styles.Alignment(horizontal='right')

    image_response = requests.get(player_image_url)
    if image_response.status_code == 200:
        image_data = BytesIO(image_response.content)
        img = Image(image_data)
        general_sheet.add_image(img, 'G3')
    else:
        print(f"Failed to download the player's image. Status code: {image_response.status_code}")
        return

    # Информация для General
    general_info = [
        "time-spent", "level", "score", "rating", "kpd", "kills", "deaths",
        "headshots", "longshots", "head-to-head-kills", "double-kills", "triple-kills",
        "games-won", "games-lost", "capture-flags"
    ]
    for i, parameter in enumerate(general_info, start=2):
        general_sheet[f'A{i}'] = parameter
        general_sheet[f'B{i}'] = stats_dict.get(parameter, 0)

    # Дополнительные листы
    sheets_data = {
        "Resources": ["ru", "gears", "skillpts", "nuclear-material"],
        "Gamemodes": [
            "games-won-domination", "games-lost-domination", "games-won-normal",
            "games-lost-normal", "games-won-assault", "games-lost-assault",
            "games-won-deathmatch", "games-won-foxhunt", "games-won-gungame"
        ],
        "Freeplay": [
            "fp-minutes-spent", "freeplay-kills", "freeplay-taken-ru", "freeplay-exit-ru",
            "freeplay-exited", "market-items-sold", "market-sold-total", "fp-valuables",
            "dog-tags-collected", "safes-opened", "enter-locked-doors",
            "consumable-item-mre-used", "consumable-item-corn-used", "consumable-item-milk-used",
            "consumable-item-radx-used", "consumable-item-beans-used", "consumable-item-paste-used",
            "consumable-item-vodka-used", "consumable-item-water-used", "consumable-item-brandy-used",
            "consumable-item-caviar-used", "consumable-item-splint-used", "consumable-item-medikit-used",
            "consumable-item-pickles-used", "consumable-item-bandages-used", "consumable-item-icecream-used",
            "consumable-item-canned-goods-used", "consumable-item-energy-drink-used",
            "consumable-item-pythonmeister-used"
        ],
        "Cases": ["cases-opened", "case-standard-opened", "case-daily-opened", "case-confiscate-opened", "case-contraband-opened"],
        "Misc": [
            "clan-kills", "clan-deaths", "clans-left", "weapon-repair", "trophies-picked",
            "disassembled-trophies", "shots", "provided-ammo", "provided-health", "events-completed"
        ]
    }

    for sheet_name, parameters in sheets_data.items():
        sheet = workbook.create_sheet(title=sheet_name)
        sheet['A1'] = "Parameter"
        sheet['B1'] = "Value"

        for i, parameter in enumerate(parameters, start=2):
            sheet[f'A{i}'] = parameter
            sheet[f'B{i}'] = stats_dict.get(parameter, 0)

    # Guns лист
    guns_sheet = workbook.create_sheet(title="Guns")
    guns_sheet['A1'] = "Weapon"
    guns_sheet['B1'] = "Kills"
    guns_sheet['C1'] = "Longshots"
    guns_sheet['D1'] = "Headshots"
    guns_sheet['E1'] = "Durability"
    guns_sheet['F1'] = "Type"  # Новая колонка "Type"
    guns_sheet['G1'] = "Class"  # Новая колонка "Class"

    # Информация для Guns

    kills_weapon_info = [key.replace("kills-from-weapon-", "") for key in stats_dict.keys() if "kills-from-weapon-" in key]
    kills_instrument_info = [key.replace("kills-from-instrument-", "") for key in stats_dict.keys() if "kills-from-instrument-" in key]
    guns_info = kills_weapon_info + kills_instrument_info

    #guns_info = [key.replace("kills-from-weapon-", "") for key in stats_dict.keys() if "kills-from-weapon-" in key]

    add_ammo_types_sheet(workbook, stats_dict)
    
    # Списки оружия для разных типов
    primary_weapons = ["galil", "mdrc", "lr300", "auga3", "sgi5k", "aswal", "groza", "aek", "gilboa-snake", "an94", "m16", "car15", "a762", "m4a1", "ls5", "rpk16", "tar21", "ft200m", "sr3", "ak105", "sig552", "aug", "hk417", "scar", "hk416", "ak12", "ak105bullpup", "ak74", "aksu", "fnfal", "akm", "ares-shrike", "scarl", "ak15", "ak74m", "famas", "aac", "ots141a", "vhs2", "acr", "m16a3", "m4a1-desert", "fn2000", "9a91", "akm74-2", "m4a1-zoom", "stg", "ar15", "ar57", "sr47", "ar18", "imbel",  "qbz", "g36c", "ak108", "cima", "ak9", "pdrc", "p91"]
    
    smg_weapons = ["vityaz", "ump", "mp5", "kiparis", "bizon", "mp5sd6", "p90", "mac10", "thompson", "augpara", "mp9", "agram", "scorpion-evo", "uzi", "pp27", "pp2000", "mp28", "m3gg", "mp40", "kriss", "mp7a2", "mpx", "mp7", "veresk"]
    
    shotgun_weapons = ["six12", "mossberg590", "aa12", "spas12", "toz34", "saiga", "spr220", "protecta", "stoeger", "mts255", "1887", "sauer", "mts569", "qbu88", "nova", "keltec", "ithaca", "qbs", "mp94", "m1014", "m1897"]
    
    # pp_weapons = ["m60", "hunter-bow", "rpk16", "rg6", "crossbow", "mg42", "pkm", "fnminimi", "m79", "mg36", "ppsh", "tul1", "pecheneg", "aek999", "rpg7", "gm94", "paint", "katana", "water", "l86", "roks3", "rpd", "rpk74", "chainsaw", "mg34", "serf", "piolet", "snowball-blaster"]
    
    sniper_weapons = ["dvl10", "kac", "vssk", "vss", "sniper", "g3", "wa2000", "sv98", "mosin", "m98b", "svu", "m98k", "garand", "vssm", "vsk94", "svds", "rsaas", "scout", "zastava", "dtsrs", "hecate", "lisle", "sks", "fg42", "gewehr43", "m1", "m14", "m21", "m82", "svd", "m110", "hksl8", "rem7600", "m95"]

    other_weapons = ["m60", "hunter-bow", "rpk16", "rg6", "crossbow", "mg42", "pkm", "fnminimi", "m79", "mg36", "ppsh", "tul1", "pecheneg", "aek999", "rpg7", "gm94", "paint", "katana", "water", "l86", "roks3", "rpd", "rpk74", "chainsaw", "mg34", "serf", "piolet", "snowball-blaster", "bar", "m134", "hk21", "mosinobrez", "m60", "hunter-bow", "rpk16", "rg6", "crossbow", "mg42", "pkm", "fnminimi", "m79", "mg36", "ppsh", "tul1", "pecheneg", "aek999", "rpg7", "gm94", "paint", "katana", "water", "l86", "roks3", "rpd", "rpk74", "chainsaw", "mg34", "serf", "piolet", "snowball-blaster"]




    secondary_weapons = ["desert50ae", "flare", "m98kobrez", "k23p", "browning", "mp443", "p50", "vp70", "mat49", "taurus",
                         "m93r", "r8", "colt", "mauser", "rgm40", "peacemaker", "usp", "coltds", "luger", "mateba", "tt",
                         "tec9", "aps", "fort12", "fnx45", "beretta", "cz75", "sigsauer", "mr412", "usp-match", "judge",
                         "pistol-s", "makarov", "desert", "p99", "mp5k", "supershorty", "magnum", "pistol", "baltiets",
                         "ots33", "fiveseven", "rsh12", "rhino", "msp", "nagant", "jericho", "scorpion"]

    knife_weapons = ["knife-type65", "knife-executor", "axe", "knife-guerrilla", "knife-kukri", "knife-tanto",
                     "knife-butterfly", "knife-puma", "knife-toyhammer", "knife-trench", "knife-bayonet", "shovel",
                     "lolipop", "crowbar", "knife-antiterror", "knife-ax", "knife", "knife-flick"]

    instrument_weapons = ["grenade-he-small", "grenade-he", "grenade-molotov", "grenade-t13", "m72law", "claymore", "c4"] 

    for i, gun in enumerate(guns_info, start=2):
        guns_sheet[f'A{i}'] = gun
        guns_sheet[f'B{i}'] = stats_dict.get(f"kills-from-weapon-{gun}", 0)
        guns_sheet[f'C{i}'] = stats_dict.get(f"longshots-from-weapon-{gun}", 0)
        guns_sheet[f'D{i}'] = stats_dict.get(f"headshots-from-weapon-{gun}", 0)
        guns_sheet[f'E{i}'] = stats_dict.get(f"durability-of-weapon-{gun}", 0)

        # Определение типа оружия
        weapon_type = "primary"
        weapon_class = "-"

        if gun in secondary_weapons:
            weapon_type = "secondary"
        elif gun in knife_weapons:
            weapon_type = "knife"
        elif gun in instrument_weapons:
            # Для оружия типа "instrument" используем правильные ключи
            weapon_type = "instrument"
            weapon_class = "-instrument"

            if gun in "m72law":
                guns_sheet[f'B{i}'] = stats_dict.get(f"kills-from-weapon-{gun}", 0)
                guns_sheet[f'C{i}'] = stats_dict.get(f"longshots-from-weapon-{gun}", 0)
                guns_sheet[f'D{i}'] = stats_dict.get(f"headshots-from-weapon-{gun}", 0)
            else:
                guns_sheet[f'B{i}'] = stats_dict.get(f"kills-from-instrument-{gun}", 0)
                guns_sheet[f'C{i}'] = stats_dict.get(f"longshots-from-instrument-{gun}", 0)
                guns_sheet[f'D{i}'] = stats_dict.get(f"headshots-from-instrument-{gun}", 0)



        # Определение типа и класса оружия
        if gun in smg_weapons:
            weapon_class = "smg"
        elif gun in primary_weapons:
            weapon_class = "rifle"
        elif gun in shotgun_weapons:
            weapon_class = "shotgun"
        # elif gun in pp_weapons:
        #     weapon_class = "pp"
        elif gun in sniper_weapons:
            weapon_class = "sniper"
        elif gun in other_weapons:
            weapon_class = "other"

        if weapon_type in "knife":
            weapon_class = "-knife"
        elif weapon_type in "secondary":
            weapon_class = "-secondary"

        guns_sheet[f'F{i}'] = weapon_type
        guns_sheet[f'G{i}'] = weapon_class


    # Сортировка листа Guns по количеству убийств
    # guns_sheet.sort(column='B', descending=True)

    # Устанавливаем ширину колонок
    for sheet in workbook.sheetnames:
        for column in workbook[sheet].columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max(10, (max_length + 2))  # Set a minimum width of 10 for empty columns
            workbook[sheet].column_dimensions[column[0].column_letter].width = adjusted_width
    # Санитизация имени файла
    sanitized_player_name = re.sub(r'[\/:*?"<>|]', '_', player_name)



    clean_gun_names(workbook)
    sort_guns_by_kills(workbook)
    add_filters_to_guns(workbook)
    clean_ammo_names(workbook)
    sort_ammo_by_kills(workbook)



    # Сохраняем таблицу
    excel_filename = f"{sanitized_player_name}.xlsx"
    workbook.save(excel_filename)
    print(f"Player statistics are saved in a file {excel_filename}")

# Запрашиваем ID игрока
# player_id = 631254
player_id = input("Player's ID: ")
player_name, player_stats_str, player_image_url = get_player_stats(player_id)

if player_stats_str:
    stats_dict = parse_stats(player_stats_str)
    create_excel_table(stats_dict, player_name, player_image_url, "player_stats")
    input("\n\n\nPress Enter to exit...")
else:
    print("Failed to get player stats.")
    input("\n\n\nPress Enter to exit...")